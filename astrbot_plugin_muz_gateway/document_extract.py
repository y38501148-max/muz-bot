# ruff: noqa: UP006 -- keep the plugin importable on Python 3.8.

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Iterable, List
from xml.etree import ElementTree

MAX_DOCUMENT_BYTES = 20 * 1024 * 1024
MAX_EXTRACTED_CHARS = 12_000
MAX_ZIP_MEMBERS = 2_000
MAX_ZIP_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ZIP_MEMBER_BYTES = 10 * 1024 * 1024
MAX_ZIP_COMPRESSION_RATIO = 100
MAX_XML_BYTES = 10 * 1024 * 1024
_SENSITIVE_PATTERNS = (
    re.compile(
        r"-----BEGIN [^-]*PRIVATE KEY-----.*?"
        r"-----END [^-]*PRIVATE KEY-----",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(r"\beyJ[A-Za-z0-9_-]{8,}\.[A-Za-z0-9_-]{8,}\.[A-Za-z0-9_-]{8,}\b"),
    re.compile(r"\bsk-[A-Za-z0-9_-]{16,}\b", re.IGNORECASE),
    re.compile(
        r"(?i)\b(?:api[_ -]?key|access[_ -]?token|secret|password|passwd)"
        r"\s*[:=]\s*[\"']?[^\s,\"';]{8,}"
    ),
    re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE),
    re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)"),
    re.compile(r"(?<!\d)\d{6}(?:18|19|20)\d{2}[01]\d[0-3]\d\d{3}[\dXx](?!\d)"),
    re.compile(r"(?<!\d)(?:\d[ -]?){13,18}\d(?!\d)"),
)
TEXT_SUFFIXES = frozenset(
    {
        ".cfg",
        ".conf",
        ".csv",
        ".html",
        ".ini",
        ".json",
        ".log",
        ".md",
        ".py",
        ".toml",
        ".tsv",
        ".txt",
        ".xml",
        ".yaml",
        ".yml",
    }
)


def _clean_text(value: object, max_chars: int) -> str:
    text = str(value or "").replace("\x00", "")
    lines = [" ".join(line.split()) for line in text.splitlines()]
    return "\n".join(line for line in lines if line)[:max_chars]


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("文件文本编码无法识别")


def _safe_zip(path: Path) -> zipfile.ZipFile:
    try:
        archive = zipfile.ZipFile(path)
    except (OSError, zipfile.BadZipFile) as error:
        raise ValueError("Office 文件结构无效") from error
    members = archive.infolist()
    total_size = sum(member.file_size for member in members)
    if (
        len(members) > MAX_ZIP_MEMBERS
        or total_size > MAX_ZIP_UNCOMPRESSED_BYTES
    ):
        archive.close()
        raise ValueError("Office 文件解压后过大")
    for member in members:
        if member.flag_bits & 0x1:
            archive.close()
            raise ValueError("不支持加密的 Office 文件")
        if member.file_size > MAX_ZIP_MEMBER_BYTES:
            archive.close()
            raise ValueError("Office 文件成员解压后过大")
        if (
            member.file_size > 0
            and member.file_size
            > max(member.compress_size, 1) * MAX_ZIP_COMPRESSION_RATIO
        ):
            archive.close()
            raise ValueError("Office 文件压缩比异常")
    return archive


def _xml_text(content: bytes) -> str:
    if len(content) > MAX_XML_BYTES:
        raise ValueError("Office XML 内容过大")
    try:
        root = ElementTree.fromstring(content)
    except ElementTree.ParseError as error:
        raise ValueError("Office XML 内容无效") from error
    values = []
    for element in root.iter():
        local_name = element.tag.rsplit("}", 1)[-1]
        if local_name in {"t", "v"} and element.text:
            values.append(element.text)
    return "\n".join(values)


def redact_sensitive_document_text(value: object) -> str:
    """Remove common credentials and direct personal identifiers."""
    result = str(value or "")
    for pattern in _SENSITIVE_PATTERNS:
        result = pattern.sub("[已隐藏]", result)
    return result


def _natural_key(value: str) -> List[object]:
    return [
        int(part) if part.isdigit() else part
        for part in re.split(r"(\d+)", value)
    ]


def _extract_office_xml(
    path: Path,
    *,
    prefixes: Iterable[str],
) -> str:
    with _safe_zip(path) as archive:
        names = sorted(
            (
                name
                for name in archive.namelist()
                if any(name.startswith(prefix) for prefix in prefixes)
                and name.endswith(".xml")
            ),
            key=_natural_key,
        )
        values = []
        for name in names:
            values.append(_xml_text(archive.read(name)))
            if sum(len(value) for value in values) >= MAX_EXTRACTED_CHARS:
                break
    return "\n".join(value for value in values if value)


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as error:
        raise ValueError("当前环境未安装 PDF 文本解析器") from error
    try:
        reader = PdfReader(str(path))
        if reader.is_encrypted or len(reader.pages) > 100:
            raise ValueError("PDF 已加密或页数过多")
        values = []
        for page in reader.pages:
            values.append(page.extract_text() or "")
            if sum(len(value) for value in values) >= MAX_EXTRACTED_CHARS:
                break
    except ValueError:
        raise
    except Exception as error:
        raise ValueError("PDF 内容无法解析") from error
    return "\n".join(values)


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("当前环境未安装表格解析器") from error
    try:
        with _safe_zip(path):
            pass
        workbook = load_workbook(
            filename=str(path),
            read_only=True,
            data_only=True,
        )
        values = []
        cell_count = 0
        for worksheet in workbook.worksheets[:20]:
            values.append(f"[工作表：{worksheet.title}]")
            for row in worksheet.iter_rows():
                row_values = [
                    str(cell.value)
                    for cell in row
                    if cell.value is not None
                ]
                cell_count += len(row_values)
                if row_values:
                    values.append("\t".join(row_values))
                if cell_count >= 20_000:
                    break
            if cell_count >= 20_000:
                break
        workbook.close()
    except Exception as error:
        raise ValueError("表格内容无法解析") from error
    return "\n".join(values)


def extract_document_text(
    path: Path,
    original_name: object,
    *,
    max_bytes: int = MAX_DOCUMENT_BYTES,
    max_chars: int = MAX_EXTRACTED_CHARS,
) -> str:
    try:
        file_size = path.stat().st_size
    except OSError as error:
        raise ValueError("文件无法读取") from error
    if file_size < 1:
        raise ValueError("文件为空")
    if file_size > max_bytes:
        raise ValueError("文件过大")

    suffix = Path(str(original_name or "")).suffix.casefold()
    if suffix in TEXT_SUFFIXES:
        extracted = _decode_text(path.read_bytes())
        if suffix == ".html":
            extracted = re.sub(r"<[^>]+>", " ", extracted)
    elif suffix == ".pdf":
        extracted = _extract_pdf(path)
    elif suffix == ".docx":
        extracted = _extract_office_xml(
            path,
            prefixes=("word/document.xml", "word/header", "word/footer"),
        )
    elif suffix == ".pptx":
        extracted = _extract_office_xml(
            path,
            prefixes=("ppt/slides/slide", "ppt/notesSlides/notesSlide"),
        )
    elif suffix in {".xlsx", ".xlsm"}:
        extracted = _extract_xlsx(path)
    else:
        raise ValueError("暂不支持该文件类型")

    normalized = _clean_text(
        redact_sensitive_document_text(extracted),
        max_chars,
    )
    if not normalized:
        raise ValueError("文件没有可提取的文本")
    return normalized
